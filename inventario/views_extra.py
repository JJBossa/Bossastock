# Vistas adicionales para las nuevas funcionalidades
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg, F
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from .models import Producto, Categoria, HistorialCambio
from .forms import ProductoForm, CategoriaForm
from .views import es_admin_bossa

def registrar_cambio(producto, usuario, tipo_cambio, campo_modificado=None, valor_anterior=None, valor_nuevo=None, descripcion=None):
    """Registra un cambio en el historial"""
    HistorialCambio.objects.create(
        producto=producto,
        usuario=usuario,
        tipo_cambio=tipo_cambio,
        campo_modificado=campo_modificado,
        valor_anterior=str(valor_anterior) if valor_anterior else None,
        valor_nuevo=str(valor_nuevo) if valor_nuevo else None,
        descripcion=descripcion
    )

@login_required
def dashboard(request):
    """Dashboard con estadísticas para el admin"""
    if not es_admin_bossa(request.user):
        messages.error(request, 'No tienes permisos para acceder al dashboard.')
        return redirect('inicio')
    
    # Estadísticas generales
    total_productos = Producto.objects.count()
    productos_activos = Producto.objects.filter(activo=True).count()
    productos_inactivos = total_productos - productos_activos
    productos_stock_bajo = Producto.objects.filter(activo=True).filter(stock__lte=F('stock_minimo')).count()
    productos_sin_imagen = Producto.objects.filter(imagen__isnull=True).count()
    
    # Valor del inventario
    valor_inventario = Producto.objects.aggregate(
        total=Sum(F('precio') * F('stock'))
    )['total'] or 0
    
    # Productos por categoría
    productos_por_categoria = Categoria.objects.annotate(
        cantidad=Count('producto')
    ).order_by('-cantidad')[:10]
    
    # Productos con stock bajo
    productos_bajo_stock = Producto.objects.filter(
        activo=True,
        stock__lte=F('stock_minimo')
    ).order_by('stock')[:10]
    
    # Productos recientes
    productos_recientes = Producto.objects.order_by('-fecha_creacion')[:5]
    
    # Cambios recientes
    cambios_recientes = HistorialCambio.objects.select_related('producto', 'usuario').order_by('-fecha')[:10]
    
    # Estadísticas de tiempo
    hoy = timezone.now().date()
    productos_hoy = Producto.objects.filter(fecha_creacion__date=hoy).count()
    productos_semana = Producto.objects.filter(fecha_creacion__gte=hoy - timedelta(days=7)).count()
    productos_mes = Producto.objects.filter(fecha_creacion__gte=hoy - timedelta(days=30)).count()
    
    context = {
        'total_productos': total_productos,
        'productos_activos': productos_activos,
        'productos_inactivos': productos_inactivos,
        'productos_stock_bajo': productos_stock_bajo,
        'productos_sin_imagen': productos_sin_imagen,
        'valor_inventario': valor_inventario,
        'productos_por_categoria': productos_por_categoria,
        'productos_bajo_stock': productos_bajo_stock,
        'productos_recientes': productos_recientes,
        'cambios_recientes': cambios_recientes,
        'productos_hoy': productos_hoy,
        'productos_semana': productos_semana,
        'productos_mes': productos_mes,
        'es_admin': True,
    }
    
    return render(request, 'inventario/dashboard.html', context)

@login_required
def detalle_producto(request, producto_id):
    """Vista de detalles de un producto"""
    producto = get_object_or_404(Producto, id=producto_id)
    historial = HistorialCambio.objects.filter(producto=producto).order_by('-fecha')[:10]
    
    context = {
        'producto': producto,
        'historial': historial,
        'es_admin': es_admin_bossa(request.user),
    }
    
    return render(request, 'inventario/detalle_producto.html', context)

@login_required
def exportar_excel(request):
    """Exporta productos a Excel"""
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Encabezados
    headers = ['SKU', 'Nombre', 'Categoría', 'Precio', 'Stock', 'Stock Mínimo', 'Valor Inventario', 'Descripción']
    ws.append(headers)
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos
    for producto in productos:
        ws.append([
            producto.sku or '',
            producto.nombre,
            producto.categoria.nombre if producto.categoria else '',
            producto.precio,
            producto.stock,
            producto.stock_minimo,
            producto.valor_inventario,
            producto.descripcion or ''
        ])
    
    # Ajustar ancho de columnas
    column_widths = [15, 30, 20, 12, 10, 12, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    
    wb.save(response)
    return response

@login_required
def exportar_pdf(request):
    """Exporta productos a PDF"""
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=1
    )
    
    # Título
    elements.append(Paragraph("Lista de Productos", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Tabla de datos
    data = [['SKU', 'Nombre', 'Categoría', 'Precio', 'Stock']]
    
    for producto in productos:
        data.append([
            producto.sku or '-',
            producto.nombre[:30],
            producto.categoria.nombre[:15] if producto.categoria else '-',
            f"${producto.precio:,.0f}",
            str(producto.stock)
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'
    return response

@login_required
def exportar_csv(request):
    """Exporta productos a CSV"""
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="productos.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['SKU', 'Nombre', 'Categoría', 'Precio', 'Stock', 'Stock Mínimo', 'Descripción'])
    
    for producto in productos:
        writer.writerow([
            producto.sku or '',
            producto.nombre,
            producto.categoria.nombre if producto.categoria else '',
            producto.precio,
            producto.stock,
            producto.stock_minimo,
            producto.descripcion or ''
        ])
    
    return response

