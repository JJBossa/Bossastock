"""
Vistas para exportación avanzada de datos
Incluye plantillas personalizables, múltiples formatos y envío por email
"""
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
import json
import csv
import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from .models import (
    Producto, Cliente, Venta, Cotizacion, 
    Factura, Proveedor, OrdenCompra, LogAccion
)
from .utils import es_admin_bossa, logger

@login_required
def exportacion_avanzada(request):
    """Vista principal para exportación avanzada"""
    if not es_admin_bossa(request.user):
        return redirect('inicio')
    
    context = {
        'es_admin': True
    }
    return render(request, 'inventario/exportacion_avanzada.html', context)

@login_required
def exportar_datos_avanzado(request):
    """Exporta datos en múltiples formatos con opciones avanzadas"""
    if not es_admin_bossa(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    
    formato = request.GET.get('formato', 'excel')
    tipo_datos = request.GET.get('tipo', 'productos')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    incluir_inactivos = request.GET.get('incluir_inactivos', 'false') == 'true'
    
    try:
        if formato == 'json':
            return exportar_json(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos)
        elif formato == 'xml':
            return exportar_xml(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos)
        elif formato == 'excel':
            return exportar_excel_avanzado(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos)
        elif formato == 'pdf':
            return exportar_pdf_avanzado(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos)
        elif formato == 'csv':
            return exportar_csv_avanzado(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos)
        else:
            return JsonResponse({'error': 'Formato no válido'}, status=400)
    except Exception as e:
        logger.error(f'Error en exportación avanzada: {str(e)}', extra={'user': request.user.username})
        return JsonResponse({'error': str(e)}, status=500)

def exportar_json(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos):
    """Exporta datos en formato JSON"""
    datos = {}
    
    if tipo_datos == 'productos':
        productos = Producto.objects.all()
        if not incluir_inactivos:
            productos = productos.filter(activo=True)
        productos = productos.select_related('categoria')
        
        datos['productos'] = [
            {
                'id': p.id,
                'nombre': p.nombre,
                'sku': p.sku,
                'categoria': p.categoria.nombre if p.categoria else None,
                'precio': float(p.precio),
                'precio_compra': float(p.precio_compra) if p.precio_compra else None,
                'stock': p.stock,
                'stock_minimo': p.stock_minimo,
                'descripcion': p.descripcion,
                'activo': p.activo,
            }
            for p in productos
        ]
    
    elif tipo_datos == 'ventas':
        ventas = Venta.objects.filter(cancelada=False).select_related('cliente', 'usuario')
        if fecha_desde:
            ventas = ventas.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta:
            ventas = ventas.filter(fecha__date__lte=fecha_hasta)
        
        datos['ventas'] = [
            {
                'id': v.id,
                'numero_venta': v.numero_venta,
                'cliente': v.cliente.nombre if v.cliente else None,
                'total': float(v.total),
                'fecha': v.fecha.isoformat(),
                'metodo_pago': v.metodo_pago,
            }
            for v in ventas
        ]
    
    response = HttpResponse(json.dumps(datos, indent=2, ensure_ascii=False), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{tipo_datos}_{datetime.now().strftime("%Y%m%d")}.json"'
    return response

def exportar_xml(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos):
    """Exporta datos en formato XML"""
    from xml.etree.ElementTree import Element, SubElement, tostring
    
    root = Element('datos')
    root.set('tipo', tipo_datos)
    root.set('fecha_exportacion', datetime.now().isoformat())
    
    if tipo_datos == 'productos':
        productos = Producto.objects.all()
        if not incluir_inactivos:
            productos = productos.filter(activo=True)
        
        productos_elem = SubElement(root, 'productos')
        for p in productos:
            prod_elem = SubElement(productos_elem, 'producto')
            SubElement(prod_elem, 'id').text = str(p.id)
            SubElement(prod_elem, 'nombre').text = p.nombre
            SubElement(prod_elem, 'sku').text = p.sku or ''
            SubElement(prod_elem, 'precio').text = str(p.precio)
            SubElement(prod_elem, 'stock').text = str(p.stock)
    
    xml_str = tostring(root, encoding='utf-8').decode('utf-8')
    response = HttpResponse(xml_str, content_type='application/xml')
    response['Content-Disposition'] = f'attachment; filename="{tipo_datos}_{datetime.now().strftime("%Y%m%d")}.xml"'
    return response

def exportar_excel_avanzado(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos):
    """Exporta datos a Excel con formato avanzado"""
    wb = Workbook()
    ws = wb.active
    ws.title = tipo_datos.title()
    
    if tipo_datos == 'productos':
        productos = Producto.objects.all()
        if not incluir_inactivos:
            productos = productos.filter(activo=True)
        productos = productos.select_related('categoria').order_by('nombre')
        
        headers = ['SKU', 'Nombre', 'Categoría', 'Precio', 'Precio Compra', 'Stock', 'Stock Mínimo', 'Valor Inventario', 'Activo']
        ws.append(headers)
        
        for producto in productos:
            ws.append([
                producto.sku or '',
                producto.nombre,
                producto.categoria.nombre if producto.categoria else '',
                producto.precio,
                producto.precio_compra or 0,
                producto.stock,
                producto.stock_minimo,
                producto.valor_inventario,
                'Sí' if producto.activo else 'No'
            ])
    
    # Estilo de encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{tipo_datos}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

def exportar_pdf_avanzado(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos):
    """Exporta datos a PDF con formato avanzado"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=20,
        alignment=1
    )
    
    elements.append(Paragraph(f"Reporte de {tipo_datos.title()}", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    if tipo_datos == 'productos':
        productos = Producto.objects.all()
        if not incluir_inactivos:
            productos = productos.filter(activo=True)
        productos = productos.select_related('categoria')
        
        data = [['SKU', 'Nombre', 'Categoría', 'Precio', 'Stock']]
        for producto in productos:
            data.append([
                producto.sku or '-',
                producto.nombre[:30],
                (producto.categoria.nombre[:15] if producto.categoria else '-'),
                f"${producto.precio:,.0f}",
                str(producto.stock)
            ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{tipo_datos}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response

def exportar_csv_avanzado(request, tipo_datos, fecha_desde, fecha_hasta, incluir_inactivos):
    """Exporta datos a CSV con opciones avanzadas"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{tipo_datos}_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    if tipo_datos == 'productos':
        productos = Producto.objects.all()
        if not incluir_inactivos:
            productos = productos.filter(activo=True)
        productos = productos.select_related('categoria')
        
        writer.writerow(['SKU', 'Nombre', 'Categoría', 'Precio', 'Stock', 'Stock Mínimo'])
        for producto in productos:
            writer.writerow([
                producto.sku or '',
                producto.nombre,
                producto.categoria.nombre if producto.categoria else '',
                producto.precio,
                producto.stock,
                producto.stock_minimo
            ])
    
    return response

